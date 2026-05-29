"""
═══════════════════════════════════════════════════════════════════════
  brief-mensuel-bourse · Streamlit Web App  ·  VERSION PRO
  ─────────────────────────────────────────────────────────────────────
  Le produit ultime d'analyse de marché : classement complet de +1200
  actions (consensus analystes + dividendes), indices mondiaux live,
  analyse technique pro (chandeliers + Bollinger + RSI + MACD),
  fondamental live (market cap, PER, beta...), comparateur, évolution
  mensuelle, et analyse IA via Claude.

  8 onglets :
    1. 📋 Classement complet  - 1237 actions, consensus + dividende + fourchette
    2. 🌍 Indices Globaux      - ~15 indices mondiaux live
    3. 📊 Graphiques agrégés   - scatter, treemap, histogrammes, top 20
    4. 🔍 Détail Ticker        - fondamental live + chart trading pro + Claude
    5. ⚖️ Comparateur          - 2-4 actions base 100
    6. 📈 Évolution            - entrées/sorties + variation de rang
    7. 📂 Par Secteur          - meilleur PEA vs CTO
    8. ℹ️ À propos

  Déploiement : Streamlit Community Cloud
    URL    : https://brief-mensuel-bourse.streamlit.app
    Repo   : https://github.com/Romtaug/LinkedIn-Brief-Mensuel-Bourse
    Source : GitHub Releases latest → ranking_latest.xlsx

  Auteur : Romain Taugourdeau
═══════════════════════════════════════════════════════════════════════
"""
from __future__ import annotations

import json
from datetime import datetime
from io import BytesIO

import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import requests
import streamlit as st
import yfinance as yf
from plotly.subplots import make_subplots


# ═════════════════════════════════════════════════════════════════════
#  CONFIG
# ═════════════════════════════════════════════════════════════════════

REPO_OWNER = "Romtaug"
REPO_NAME  = "LinkedIn-Brief-Mensuel-Bourse"
XLSX_URL   = f"https://github.com/{REPO_OWNER}/{REPO_NAME}/releases/latest/download/ranking_latest.xlsx"
RAW_BASE   = f"https://raw.githubusercontent.com/{REPO_OWNER}/{REPO_NAME}/main"

COLORS = {
    "bg":       "#0a1628",
    "bg_panel": "#0f1d34",
    "bg_row":   "#13243f",
    "grid":     "#1e3358",
    "blue":     "#00bfe6",
    "blue_dim": "#0080a0",
    "amber":    "#ffaa00",
    "text":     "#e8eaf0",
    "text_mid": "#8b9bb4",
    "dim":      "#4a5c7a",
    "green":    "#00ff66",
    "red":      "#ff2e2e",
    "gold":     "#ffd700",
    "purple":   "#b47cff",
    "pea":      "#00d4ff",
    "cto":      "#7c9eff",
}

PARRAINAGE_URL = "https://bour.so/p/GB93ZfQVNVr"
CODE_PARRAIN   = "ROTA0058"
LINKEDIN_URL   = "https://www.linkedin.com/in/romain-taugourdeau/"

# Indices mondiaux pour l'onglet "Indices Globaux"
GLOBAL_INDICES = {
    "^GSPC":     ("🇺🇸 S&P 500",        "Actions US large cap"),
    "^NDX":      ("🇺🇸 NASDAQ 100",     "Tech US"),
    "^DJI":      ("🇺🇸 Dow Jones",      "Industrielles US"),
    "^STOXX":    ("🇪🇺 STOXX 600",      "Actions européennes"),
    "^STOXX50E": ("🇪🇺 Euro Stoxx 50",  "50 plus grandes capi zone euro"),
    "^FCHI":     ("🇫🇷 CAC 40",         "Actions françaises"),
    "^GDAXI":    ("🇩🇪 DAX 40",         "Actions allemandes"),
    "^FTSE":     ("🇬🇧 FTSE 100",       "Actions britanniques"),
    "^IBEX":     ("🇪🇸 IBEX 35",        "Actions espagnoles"),
    "^N225":     ("🇯🇵 Nikkei 225",     "Actions japonaises"),
    "^HSI":      ("🇭🇰 Hang Seng",      "Actions Hong Kong"),
    "^VIX":      ("😱 VIX",             "Indice de la peur (volatilité)"),
    "GC=F":      ("🥇 Or (Gold)",       "Once d'or USD"),
    "CL=F":      ("🛢️ Pétrole WTI",     "Baril WTI USD"),
    "EURUSD=X":  ("💱 EUR/USD",         "Taux de change"),
    "BTC-EUR":   ("₿ Bitcoin",          "Crypto BTC en EUR"),
}


# ═════════════════════════════════════════════════════════════════════
#  PAGE CONFIG + CSS
# ═════════════════════════════════════════════════════════════════════

st.set_page_config(
    page_title="Brief Mensuel Bourse",
    page_icon="📈",
    layout="wide",
    initial_sidebar_state="expanded",
    menu_items={
        "Get Help": LINKEDIN_URL,
        "About": "Brief Mensuel Bourse - Analyse automatisée de +1200 actions PEA & CTO.",
    },
)

st.markdown(f"""
<style>
    /* ── Forçage du thème dark (indépendant de config.toml) ── */
    .stApp {{
        background: {COLORS['bg']} !important;
        color: {COLORS['text']} !important;
    }}
    [data-testid="stHeader"] {{
        background: {COLORS['bg']} !important;
    }}
    [data-testid="stAppViewContainer"], .main, .block-container {{
        background: {COLORS['bg']} !important;
        color: {COLORS['text']} !important;
    }}
    /* Widgets natifs (selectbox, slider, radio, multiselect) sur fond sombre */
    [data-baseweb="select"] > div, [data-baseweb="input"] > div {{
        background: {COLORS['bg_panel']} !important;
        border-color: {COLORS['grid']} !important;
    }}
    .stSlider [data-baseweb="slider"] {{ color: {COLORS['blue']} !important; }}
    label, .stMarkdown, p, span, li {{ color: {COLORS['text']}; }}
    [data-testid="stMetricLabel"] {{ color: {COLORS['text_mid']} !important; }}

    html, body, [class*="css"] {{
        font-family: 'JetBrains Mono', 'Courier New', monospace;
    }}
    h1, h2, h3, h4 {{
        font-family: 'Inter', sans-serif !important;
        letter-spacing: -0.5px;
        color: {COLORS['text']};
    }}
    .bar-top {{
        background: {COLORS['blue']};
        color: {COLORS['bg']};
        padding: 12px 24px;
        font-weight: 700;
        font-size: 14px;
        letter-spacing: 1px;
        margin: -2rem -2rem 1.5rem -2rem;
        display: flex;
        align-items: center;
        justify-content: space-between;
    }}
    .live-dot {{
        display: inline-block;
        width: 10px; height: 10px;
        background: {COLORS['bg']};
        border-radius: 50%;
        margin-right: 12px;
        animation: pulse 2s infinite;
    }}
    @keyframes pulse {{
        0%, 100% {{ opacity: 1; }}
        50% {{ opacity: 0.4; }}
    }}
    [data-testid="stMetricValue"] {{
        font-family: 'Inter', sans-serif;
        color: {COLORS['blue']};
        font-weight: 800;
    }}
    [data-testid="stSidebar"] {{
        background: {COLORS['bg_panel']};
    }}
    .stButton > button {{
        background: {COLORS['bg_panel']};
        border: 2px solid {COLORS['blue']};
        color: {COLORS['blue']};
        font-weight: 700;
        letter-spacing: 1px;
    }}
    .stButton > button:hover {{
        background: {COLORS['blue']};
        color: {COLORS['bg']};
    }}
    .stTabs [data-baseweb="tab-list"] {{ gap: 6px; flex-wrap: wrap; width: 100%; }}
    .stTabs [data-baseweb="tab"] {{
        background: {COLORS['bg_panel']};
        border-radius: 0;
        color: {COLORS['text_mid']};
        font-weight: 700;
        font-size: 13px;
        flex: 1 1 0;
        justify-content: center;
        text-align: center;
        white-space: nowrap;
    }}
    .stTabs [aria-selected="true"] {{
        background: {COLORS['blue']};
        color: {COLORS['bg']};
    }}
    a {{ color: {COLORS['blue']} !important; }}
    .info-card {{
        background: {COLORS['bg_panel']};
        border-left: 3px solid {COLORS['blue']};
        padding: 14px 18px;
        margin: 8px 0;
        border-radius: 4px;
    }}
    .info-card .label {{ color: {COLORS['text_mid']}; font-size: 11px; text-transform: uppercase; letter-spacing: 1px; }}
    .info-card .value {{ color: {COLORS['text']}; font-size: 22px; font-weight: 800; font-family: 'Inter', sans-serif; }}
</style>
""", unsafe_allow_html=True)


# ═════════════════════════════════════════════════════════════════════
#  DATA LOADING (cached)
# ═════════════════════════════════════════════════════════════════════

@st.cache_data(ttl=6 * 3600, show_spinner="📡 Chargement du brief mensuel...")
def load_xlsx_from_github(url: str = XLSX_URL) -> dict:
    """Télécharge l'xlsx depuis GitHub Releases (cache 6h). Dict {sheet: df}."""
    try:
        r = requests.get(url, timeout=30, headers={"User-Agent": "BriefBourseApp/2.0"})
        r.raise_for_status()
        return pd.read_excel(BytesIO(r.content), sheet_name=None)
    except requests.HTTPError as e:
        st.error(f"❌ Erreur HTTP {e.response.status_code} - la release n'existe peut-être pas encore.")
        return {}
    except Exception as e:
        st.error(f"❌ Erreur de chargement : {e}")
        return {}


@st.cache_data(ttl=3600, show_spinner=False)
def fetch_ticker_history(ticker: str, period: str = "1y") -> pd.DataFrame:
    """Fetch yfinance OHLCV (cache 1h). Garde Open/High/Low/Close/Volume."""
    try:
        hist = yf.Ticker(ticker).history(period=period, auto_adjust=True)
        if hist.empty:
            return pd.DataFrame()
        cols = [c for c in ["Open", "High", "Low", "Close", "Volume"] if c in hist.columns]
        return hist[cols].copy()
    except Exception:
        return pd.DataFrame()


@st.cache_data(ttl=3600, show_spinner=False)
def fetch_ticker_info(ticker: str) -> dict:
    """
    Fetch yfinance .info LIVE (cache 1h) → fondamental complet.
    Renvoie un dict nettoyé. Sur les indices, beaucoup de champs sont None.
    """
    try:
        info = yf.Ticker(ticker).info or {}
        return info
    except Exception:
        return {}


@st.cache_data(ttl=1800, show_spinner="🌍 Chargement des indices mondiaux...")
def fetch_global_indices() -> pd.DataFrame:
    """
    Fetch les ~15 indices mondiaux (cache 30 min).
    Calcule perfs jour / mois / an + série 1 an pour mini-graph.
    """
    rows = []
    for tk, (label, desc) in GLOBAL_INDICES.items():
        try:
            h = yf.Ticker(tk).history(period="1y", auto_adjust=True)
            if h.empty or "Close" not in h.columns:
                continue
            close = h["Close"].dropna()
            if len(close) < 2:
                continue
            last = float(close.iloc[-1])
            prev = float(close.iloc[-2])
            perf_day = (last / prev - 1) * 100
            # perf 1 mois (~21 séances)
            ref_m = close.iloc[-22] if len(close) > 22 else close.iloc[0]
            perf_month = (last / float(ref_m) - 1) * 100
            # perf 1 an
            perf_year = (last / float(close.iloc[0]) - 1) * 100
            rows.append({
                "ticker": tk,
                "label": label,
                "desc": desc,
                "last": last,
                "perf_day": perf_day,
                "perf_month": perf_month,
                "perf_year": perf_year,
                "series": close.values.tolist(),
                "dates": [d.strftime("%Y-%m-%d") for d in close.index],
            })
        except Exception:
            continue
    return pd.DataFrame(rows)


@st.cache_data(ttl=6 * 3600, show_spinner=False)
def load_snapshot(month_key: str) -> pd.DataFrame:
    """Charge un snapshot mensuel JSON depuis GitHub raw (ex: '2026-05')."""
    url = f"{RAW_BASE}/snapshots/ranking_{month_key}.json"
    try:
        r = requests.get(url, timeout=20, headers={"User-Agent": "BriefBourseApp/2.0"})
        r.raise_for_status()
        data = r.json()
        return pd.DataFrame(data) if isinstance(data, list) else pd.DataFrame()
    except Exception:
        return pd.DataFrame()


# ═════════════════════════════════════════════════════════════════════
#  INDICATEURS TECHNIQUES
# ═════════════════════════════════════════════════════════════════════

def compute_rsi(close: pd.Series, period: int = 14) -> pd.Series:
    """RSI de Wilder (lissage exponentiel alpha=1/period)."""
    delta = close.diff()
    gain = delta.clip(lower=0)
    loss = -delta.clip(upper=0)
    avg_gain = gain.ewm(alpha=1 / period, min_periods=period, adjust=False).mean()
    avg_loss = loss.ewm(alpha=1 / period, min_periods=period, adjust=False).mean()
    rs = avg_gain / avg_loss
    return 100 - (100 / (1 + rs))


def compute_macd(close: pd.Series, fast: int = 12, slow: int = 26, signal: int = 9):
    """MACD = EMA12 - EMA26 ; signal = EMA9(MACD) ; hist = MACD - signal."""
    ema_fast = close.ewm(span=fast, adjust=False).mean()
    ema_slow = close.ewm(span=slow, adjust=False).mean()
    macd = ema_fast - ema_slow
    sig = macd.ewm(span=signal, adjust=False).mean()
    return macd, sig, macd - sig


def compute_bollinger(close: pd.Series, window: int = 20, n_std: float = 2.0):
    """Bandes de Bollinger : MA20 ± 2σ."""
    ma = close.rolling(window).mean()
    std = close.rolling(window).std()
    return ma, ma + n_std * std, ma - n_std * std


# ═════════════════════════════════════════════════════════════════════
#  HELPERS
# ═════════════════════════════════════════════════════════════════════

def get_country_from_ticker(ticker: str) -> str:
    flag_map = {
        ".PA": "🇫🇷 France", ".DE": "🇩🇪 Allemagne", ".AS": "🇳🇱 Pays-Bas",
        ".BR": "🇧🇪 Belgique", ".MI": "🇮🇹 Italie", ".MC": "🇪🇸 Espagne",
        ".LS": "🇵🇹 Portugal", ".OL": "🇳🇴 Norvège", ".ST": "🇸🇪 Suède",
        ".HE": "🇫🇮 Finlande", ".CO": "🇩🇰 Danemark", ".VI": "🇦🇹 Autriche",
        ".IR": "🇮🇪 Irlande", ".SW": "🇨🇭 Suisse", ".L": "🇬🇧 Royaume-Uni",
        ".WA": "🇵🇱 Pologne", ".AT": "🇬🇷 Grèce", ".T": "🇯🇵 Japon",
        ".TO": "🇨🇦 Canada", ".AX": "🇦🇺 Australie", ".HK": "🇭🇰 Hong Kong",
    }
    if not isinstance(ticker, str) or "." not in ticker:
        return "🇺🇸 États-Unis"
    suf = "." + ticker.rsplit(".", 1)[1]
    return flag_map.get(suf, "🌍 Autre")


def fmt_large_number(n, currency: str = "€") -> str:
    """Formate un grand nombre : 2.4T€, 850.3B€, 12.5M€."""
    if n is None or (isinstance(n, float) and np.isnan(n)):
        return "-"
    try:
        n = float(n)
    except (TypeError, ValueError):
        return "-"
    sign = "-" if n < 0 else ""
    n = abs(n)
    if n >= 1e12:
        return f"{sign}{n/1e12:.2f} T{currency}"
    if n >= 1e9:
        return f"{sign}{n/1e9:.1f} Md{currency}"
    if n >= 1e6:
        return f"{sign}{n/1e6:.1f} M{currency}"
    if n >= 1e3:
        return f"{sign}{n/1e3:.1f} k{currency}"
    return f"{sign}{n:.0f} {currency}"


def reco_to_stars(reco_mean) -> str:
    """reco_mean yfinance : 1=Strong Buy ... 5=Strong Sell. → étoiles."""
    if reco_mean is None or (isinstance(reco_mean, float) and np.isnan(reco_mean)):
        return "-"
    try:
        score = 6 - float(reco_mean)  # 1→5 étoiles, 5→1 étoile
    except (TypeError, ValueError):
        return "-"
    full = int(round(score))
    full = max(0, min(5, full))
    return "★" * full + "☆" * (5 - full)


def fmt_pct(v, plus: bool = True) -> str:
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return "-"
    try:
        v = float(v)
    except (TypeError, ValueError):
        return "-"
    return f"{v:+.1f}%" if plus else f"{v:.2f}%"


def color_for_pct(v) -> str:
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return COLORS["text_mid"]
    return COLORS["green"] if v >= 0 else COLORS["red"]


def trend_signal(row: dict) -> tuple[str, str]:
    """
    Signal de tendance SIMPLE pour le tableau (momentum + consensus + potentiel).
    ⚠️ Ce n'est PAS un signal technique complet (RSI/MACD/MA) — celui-là est
    dans l'onglet Détail Ticker. Ici on combine les données déjà dispo dans l'Excel.
    Retourne (emoji, label).
    """
    perf = row.get("perf_1m")
    reco = row.get("reco_mean")   # 1 = achat fort ... 5 = vente forte
    pot  = row.get("total_pct")
    score = 0
    if perf is not None and pd.notna(perf):
        score += 1 if perf > 2 else (-1 if perf < -2 else 0)
    if reco is not None and pd.notna(reco):
        score += 1 if reco <= 2.2 else (-1 if reco >= 3.2 else 0)
    if pot is not None and pd.notna(pot):
        score += 1 if pot > 10 else (-1 if pot < 0 else 0)
    if score >= 2:
        return ("🟢", "Haussier")
    if score <= -1:
        return ("🔴", "Prudence")
    return ("🟡", "Neutre")


def build_claude_prompt(row: dict, info: dict | None = None) -> str:
    """Prompt Claude PRO : recherche web forcée + analyse détaillée + notation multi-critères."""
    pea_str = "Oui ✅ (éligible PEA)" if row.get("pea") else "Non ❌ (CTO uniquement)"
    country = get_country_from_ticker(row["ticker"])
    div = row.get("div_pct", 0) or 0
    div_str = f"{div:.2f}%" if div else "0% (ne verse pas de dividende)"

    # ── Fondamental live enrichi ─────────────────────────────────────
    fond = ""
    if info:
        mc   = info.get("marketCap")
        pe   = info.get("trailingPE")
        fpe  = info.get("forwardPE")
        beta = info.get("beta")
        pb   = info.get("priceToBook")
        w52h = info.get("fiftyTwoWeekHigh")
        w52l = info.get("fiftyTwoWeekLow")
        emp  = info.get("fullTimeEmployees")
        ind  = info.get("industry")
        if mc:   fond += f"\nCapitalisation boursière : {fmt_large_number(mc, '')}"
        if pe:   fond += f"\nPER (trailing) : {pe:.1f}"
        if fpe:  fond += f"\nPER (forward / anticipé) : {fpe:.1f}"
        if pb:   fond += f"\nPrice / Book : {pb:.2f}"
        if beta: fond += f"\nBeta (volatilité vs marché) : {beta:.2f}"
        if w52h: fond += f"\nPlus haut 52 sem. : {w52h:.2f} {row.get('currency','')}"
        if w52l: fond += f"\nPlus bas 52 sem. : {w52l:.2f} {row.get('currency','')}"
        if ind:  fond += f"\nIndustrie : {ind}"
        if emp:  fond += f"\nEffectif : {emp:,} employés".replace(",", " ")
    if not fond:
        fond = "\n(Fondamental live indisponible - à compléter par ta recherche web.)"

    return f"""Tu es le meilleur analyste financier du marché, indépendant et rigoureux, spécialisé en actions cotées. Tu produis une analyse d'investissement complète, chiffrée et sans complaisance.

⚠️ IMPORTANT - UTILISE TES OUTILS DE RECHERCHE WEB. Les données ci-dessous sont une photo figée. AVANT de rédiger, recherche sur le web les informations à jour suivantes :
• Les DERNIERS résultats trimestriels/annuels publiés (chiffre d'affaires, BPA, marges, guidance)
• Les actualités récentes (3-6 derniers mois) : annonces, contrats, rachats, litiges, changements de direction
• Les révisions récentes de recommandations/objectifs d'analystes
• La position concurrentielle et les tendances de son secteur
• Tout risque réglementaire, macro ou spécifique récent
Croise systématiquement ces données fraîches avec les chiffres figés ci-dessous (ils peuvent dater).

═══════════════════════════════════
DONNÉES (photo figée - à actualiser par recherche)
═══════════════════════════════════
Société : {row.get('name', '')} ({row['ticker']})
Secteur : {row.get('sector_fr', '')}
Pays / Marché : {country}
Éligibilité PEA : {pea_str}
─────
Cours : {row.get('price_eur', '?')}€ (natif {row.get('price', '?')} {row.get('currency', '')})
Consensus analystes : cible 12 mois {fmt_pct(row.get('target_pct'))} · {row.get('reco_label', '-')} · {int(row.get('analyst_count', 0) or 0)} analystes
Fourchette des objectifs : bas {fmt_pct(row.get('target_low_pct'))} / haut {fmt_pct(row.get('target_high_pct'))}
Dividende : {div_str}
Performance mois précédent : {fmt_pct(row.get('perf_1m'))}
Potentiel total (cible + dividende) : {fmt_pct(row.get('total_pct'))}{fond}

═══════════════════════════════════
STRUCTURE DE TA RÉPONSE (en français, détaillée)
═══════════════════════════════════
**1. L'entreprise** - Activité, modèle économique, marchés, position concurrentielle, avantages compétitifs (moat). 4-6 lignes.

**2. Actualité récente** (issue de ta recherche web) - Derniers résultats publiés, guidance, news marquantes, dynamique récente. Cite les chiffres clés et leur date.

**3. Lecture des chiffres** - Valorisation (PER/PB vs secteur et historique : cher ou bon marché ?), croissance, rentabilité/marges, santé financière (dette, cash-flow), momentum, ce que dit (et ne dit pas) le consensus.

**4. Thèse haussière vs baissière** - 3 arguments solides de chaque côté, hiérarchisés par importance.

**5. NOTATION MULTI-CRITÈRES** - Note chaque critère de 0 à 10 et justifie en 1 ligne :
   • 💰 Valorisation (cher/pas cher vs valeur réelle)
   • 📈 Croissance (CA, BPA, perspectives)
   • 🏆 Qualité / rentabilité (marges, ROE, moat)
   • 🩺 Santé financière (dette, trésorerie, cash-flow)
   • ⚡ Momentum (tendance de cours + dynamique récente)
   • 👥 Consensus analystes (force et fiabilité)
   • 💶 Rendement (dividende + politique de retour)
   • ⚠️ Risque (inversé : 10 = peu risqué)
   Puis donne une **NOTE GLOBALE PONDÉRÉE /10** avec une phrase de synthèse.

**6. Profil & risques** - À quel type d'investisseur ce titre convient (prudent / équilibré / dynamique ; horizon ; PEA ou CTO) ? Quels sont les 2-3 risques majeurs à surveiller ?

**7. Conclusion** - Synthèse actionnable en 3-4 lignes : l'essentiel à retenir.

⚠️ Termine en précisant clairement que ceci n'est PAS un conseil en investissement, que l'analyse peut contenir des erreurs, et que tout investissement comporte un risque de perte en capital. L'investisseur doit faire ses propres recherches."""


def plot_styled_layout(fig, title: str = "", height: int = 400):
    fig.update_layout(
        title=title,
        paper_bgcolor=COLORS["bg"],
        plot_bgcolor=COLORS["bg_panel"],
        font=dict(family="JetBrains Mono, monospace", color=COLORS["text"], size=12),
        title_font=dict(family="Inter, sans-serif", size=18, color=COLORS["text"]),
        xaxis=dict(gridcolor=COLORS["grid"], linecolor=COLORS["grid"], zeroline=False),
        yaxis=dict(gridcolor=COLORS["grid"], linecolor=COLORS["grid"], zeroline=False),
        legend=dict(bgcolor=COLORS["bg_panel"], bordercolor=COLORS["grid"], borderwidth=1),
        height=height,
        margin=dict(l=40, r=40, t=60, b=40),
    )
    return fig


def _write_styled_sheet(ws, df_export: pd.DataFrame, pct_cols: list, price_cols: list,
                        color_scale_cols: list | None = None):
    """Écrit un DataFrame dans une feuille stylée (en-têtes Bloomberg, couleurs %, etc.)."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule

    headers = list(df_export.columns)
    header_fill = PatternFill("solid", fgColor="0A1628")
    header_font = Font(bold=True, color="00D4FF", size=11, name="Calibri")
    alt_fill = PatternFill("solid", fgColor="F2F6FC")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    thin = Side(style="thin", color="D6DEE8")
    border = Border(bottom=thin)

    pct_fmt = '[Green]+0.0"%";[Red]-0.0"%";0.0"%"'
    price_fmt = '#,##0.00" €"'

    # En-têtes
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    # Données
    for r, (_, row) in enumerate(df_export.iterrows(), 2):
        for c, h in enumerate(headers, 1):
            val = row[h]
            if isinstance(val, float) and pd.isna(val):
                val = None
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = border
            if r % 2 == 1:
                cell.fill = alt_fill
            if h in pct_cols and isinstance(val, (int, float)):
                cell.number_format = pct_fmt
                cell.alignment = center
            elif h in price_cols and isinstance(val, (int, float)):
                cell.number_format = price_fmt
                cell.alignment = center
            elif isinstance(val, (int, float)):
                cell.alignment = center
            else:
                cell.alignment = left

    n_rows = len(df_export)

    # Mise en forme conditionnelle (échelle rouge→jaune→vert) sur colonnes choisies
    if color_scale_cols and n_rows > 0:
        for h in color_scale_cols:
            if h in headers:
                ci = headers.index(h) + 1
                col = get_column_letter(ci)
                rng = f"{col}2:{col}{n_rows + 1}"
                ws.conditional_formatting.add(rng, ColorScaleRule(
                    start_type="percentile", start_value=5,  start_color="FFF8696B",
                    mid_type="percentile",   mid_value=50,   mid_color="FFFFEB84",
                    end_type="percentile",   end_value=95,   end_color="FF63BE7B",
                ))

    # Largeurs colonnes
    for c, h in enumerate(headers, 1):
        max_len = len(str(h))
        for v in df_export[h].head(200):
            max_len = max(max_len, len(str(v)) if v is not None else 0)
        ws.column_dimensions[get_column_letter(c)].width = min(max(max_len + 5, 10), 45)

    ws.freeze_panes = "A2"
    if n_rows > 0:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{n_rows + 1}"


def _write_synthesis_sheet(ws, kpis: dict, top_df: pd.DataFrame, pct_cols: list, price_cols: list):
    """Première feuille : bandeau titre + KPIs + top opportunités."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    title_fill = PatternFill("solid", fgColor="0A1628")
    title_font = Font(bold=True, color="00D4FF", size=16, name="Calibri")
    sub_font = Font(color="8B9BB4", size=10, name="Calibri", italic=True)
    kpi_label_font = Font(color="4A5C7A", size=10, name="Calibri", bold=True)
    kpi_val_font = Font(color="0A1628", size=14, name="Calibri", bold=True)

    # Bandeau titre (lignes 1-2)
    ws.merge_cells("A1:F1")
    c = ws.cell(row=1, column=1, value="BRIEF MENSUEL BOURSE — Synthèse")
    c.fill = title_fill; c.font = title_font
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A2:F2")
    c2 = ws.cell(row=2, column=1,
                 value=f"Généré le {datetime.now():%d/%m/%Y à %H:%M} · Données Yahoo Finance · "
                       f"Romain Taugourdeau")
    c2.font = sub_font

    # KPIs (lignes 4+)
    r = 4
    for label, value in kpis.items():
        lc = ws.cell(row=r, column=1, value=label); lc.font = kpi_label_font
        vc = ws.cell(row=r, column=2, value=value); vc.font = kpi_val_font
        r += 1

    # Top opportunités (titre + tableau)
    r += 1
    th = ws.cell(row=r, column=1, value="🏆 TOP OPPORTUNITÉS (par potentiel total)")
    th.font = Font(bold=True, color="0A1628", size=12, name="Calibri")
    r += 1
    _write_styled_sheet_at(ws, top_df, pct_cols, price_cols, start_row=r,
                           color_scale_cols=["Potentiel"])

    ws.column_dimensions["A"].width = 26
    for col in ["B", "C", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[col].width = max(ws.column_dimensions[col].width or 0, 14)


def _write_styled_sheet_at(ws, df_export, pct_cols, price_cols, start_row=1, color_scale_cols=None):
    """Comme _write_styled_sheet mais à partir d'une ligne donnée (pour la feuille Synthèse)."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule

    headers = list(df_export.columns)
    header_fill = PatternFill("solid", fgColor="0A1628")
    header_font = Font(bold=True, color="00D4FF", size=11, name="Calibri")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    pct_fmt = '[Green]+0.0"%";[Red]-0.0"%";0.0"%"'
    price_fmt = '#,##0.00" €"'

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=c, value=h)
        cell.fill = header_fill; cell.font = header_font; cell.alignment = center

    for i, (_, row) in enumerate(df_export.iterrows(), 1):
        rr = start_row + i
        for c, h in enumerate(headers, 1):
            val = row[h]
            if isinstance(val, float) and pd.isna(val):
                val = None
            cell = ws.cell(row=rr, column=c, value=val)
            if h in pct_cols and isinstance(val, (int, float)):
                cell.number_format = pct_fmt; cell.alignment = center
            elif h in price_cols and isinstance(val, (int, float)):
                cell.number_format = price_fmt; cell.alignment = center
            elif isinstance(val, (int, float)):
                cell.alignment = center
            else:
                cell.alignment = left

    n = len(df_export)
    if color_scale_cols and n > 0:
        for h in color_scale_cols:
            if h in headers:
                col = get_column_letter(headers.index(h) + 1)
                rng = f"{col}{start_row+1}:{col}{start_row+n}"
                ws.conditional_formatting.add(rng, ColorScaleRule(
                    start_type="percentile", start_value=5,  start_color="FFF8696B",
                    mid_type="percentile",   mid_value=50,   mid_color="FFFFEB84",
                    end_type="percentile",   end_value=95,   end_color="FF63BE7B",
                ))


def make_workbook_excel(sheets: dict, pct_cols: list, price_cols: list,
                        synth: dict | None = None,
                        color_scale_cols: list | None = None) -> bytes:
    """
    Génère un classeur Excel multi-feuilles.
    sheets : {nom_feuille: DataFrame}  ·  synth : {kpis, top_df} optionnel (1re feuille).
    """
    from openpyxl import Workbook

    wb = Workbook()
    default = wb.active
    first_used = False

    if synth:
        default.title = "Synthèse"
        _write_synthesis_sheet(default, synth["kpis"], synth["top_df"], pct_cols, price_cols)
        first_used = True

    for name, dfx in sheets.items():
        if dfx is None or len(dfx) == 0:
            continue
        if not first_used:
            ws = default; ws.title = name[:31]; first_used = True
        else:
            ws = wb.create_sheet(name[:31])
        _write_styled_sheet(ws, dfx, pct_cols, price_cols, color_scale_cols=color_scale_cols)

    if not first_used:  # sécurité : aucun onglet écrit
        default.title = "Vide"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def make_styled_excel(df_export: pd.DataFrame, pct_cols: list, price_cols: list,
                      sheet_name: str = "Classement") -> bytes:
    """Compat : Excel mono-feuille (utilise le helper de feuille stylée)."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    _write_styled_sheet(ws, df_export, pct_cols, price_cols, color_scale_cols=["Potentiel"])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_technical_summary(hist: pd.DataFrame) -> dict:
    """
    Analyse les indicateurs techniques et génère une lecture en langage naturel.
    100% descriptif/factuel - JAMAIS un conseil d'achat/vente.

    Returns dict:
      - verdict   : phrase de synthèse globale
      - bias      : "haussier" | "baissier" | "neutre"
      - bias_score: int (somme des signaux, négatif=baissier, positif=haussier)
      - signals   : liste de (emoji, titre, texte) par indicateur
    """
    close = hist["Close"].dropna()
    if len(close) < 30:
        return {"verdict": "Pas assez d'historique pour une lecture technique fiable.",
                "bias": "neutre", "bias_score": 0, "signals": []}

    last = float(close.iloc[-1])
    signals = []
    score = 0  # >0 haussier, <0 baissier

    # ── 1. TENDANCE (cours vs moyennes mobiles) ──────────────────────
    ma20 = close.rolling(20).mean().iloc[-1]
    ma50 = close.rolling(50).mean().iloc[-1]
    ma200 = close.rolling(200).mean().iloc[-1] if len(close) >= 200 else None

    above20 = last > ma20
    above50 = last > ma50
    above200 = (ma200 is not None) and (last > ma200)

    if ma200 is not None:
        if above20 and above50 and above200:
            trend_txt = "Le cours est au-dessus de ses moyennes 20, 50 et 200 jours → tendance haussière confirmée sur tous les horizons."
            score += 2
        elif (not above20) and (not above50) and (not above200):
            trend_txt = "Le cours est sous ses moyennes 20, 50 et 200 jours → tendance baissière sur tous les horizons."
            score -= 2
        elif above50 and not above200:
            trend_txt = "Le cours repasse au-dessus de sa moyenne 50 jours mais reste sous la 200 jours → rebond de court terme dans une tendance de fond encore baissière."
            score += 0
        elif above200 and not above50:
            trend_txt = "Le cours est au-dessus de sa moyenne 200 jours mais sous la 50 jours → consolidation dans une tendance de fond haussière."
            score += 0
        else:
            trend_txt = "Position mitigée vis-à-vis des moyennes mobiles → tendance indécise (phase de transition)."
    else:
        if above20 and above50:
            trend_txt = "Le cours est au-dessus de ses moyennes 20 et 50 jours → orientation haussière de court/moyen terme (historique insuffisant pour la 200 j)."
            score += 1
        elif (not above20) and (not above50):
            trend_txt = "Le cours est sous ses moyennes 20 et 50 jours → orientation baissière de court/moyen terme."
            score -= 1
        else:
            trend_txt = "Position mitigée vis-à-vis des moyennes 20/50 jours → tendance indécise."
    signals.append(("📐", "Tendance (moyennes mobiles)", trend_txt))

    # ── 2. RSI ───────────────────────────────────────────────────────
    rsi = compute_rsi(close).iloc[-1]
    if pd.notna(rsi):
        if rsi >= 70:
            rsi_txt = f"RSI à {rsi:.0f} → zone de surachat (>70). Le titre a beaucoup monté à court terme, un essoufflement ou une respiration est possible."
            score -= 1
        elif rsi <= 30:
            rsi_txt = f"RSI à {rsi:.0f} → zone de survente (<30). Le titre a beaucoup baissé à court terme, un rebound technique est possible."
            score += 1
        elif rsi >= 55:
            rsi_txt = f"RSI à {rsi:.0f} → zone neutre à légèrement acheteuse (momentum positif sans excès)."
        elif rsi <= 45:
            rsi_txt = f"RSI à {rsi:.0f} → zone neutre à légèrement vendeuse (momentum mou)."
        else:
            rsi_txt = f"RSI à {rsi:.0f} → zone neutre (ni surachat ni survente)."
        signals.append(("⚡", "Momentum (RSI 14)", rsi_txt))

    # ── 3. MACD ──────────────────────────────────────────────────────
    macd, sig, _ = compute_macd(close)
    if len(macd) >= 2 and pd.notna(macd.iloc[-1]) and pd.notna(sig.iloc[-1]):
        m_now, s_now = macd.iloc[-1], sig.iloc[-1]
        m_prev, s_prev = macd.iloc[-2], sig.iloc[-2]
        crossed_up = (m_prev <= s_prev) and (m_now > s_now)
        crossed_dn = (m_prev >= s_prev) and (m_now < s_now)
        if crossed_up:
            macd_txt = "Croisement haussier récent du MACD (la ligne MACD repasse au-dessus de sa ligne de signal) → signal de momentum positif."
            score += 1
        elif crossed_dn:
            macd_txt = "Croisement baissier récent du MACD (la ligne MACD repasse sous sa ligne de signal) → signal de momentum négatif."
            score -= 1
        elif m_now > s_now:
            macd_txt = "MACD au-dessus de sa ligne de signal → momentum haussier en cours."
            score += 1
        else:
            macd_txt = "MACD sous sa ligne de signal → momentum baissier en cours."
            score -= 1
        signals.append(("📊", "MACD (12,26,9)", macd_txt))

    # ── 4. BOLLINGER (position dans les bandes) ──────────────────────
    bmid, bup, blo = compute_bollinger(close)
    if pd.notna(bup.iloc[-1]) and pd.notna(blo.iloc[-1]):
        up, lo, mid = bup.iloc[-1], blo.iloc[-1], bmid.iloc[-1]
        rng = up - lo
        if rng > 0:
            pos = (last - lo) / rng  # 0 = bande basse, 1 = bande haute
            if pos >= 0.95:
                boll_txt = "Le cours touche la bande supérieure de Bollinger → tension acheteuse forte (parfois précède une respiration)."
            elif pos <= 0.05:
                boll_txt = "Le cours touche la bande inférieure de Bollinger → pression vendeuse forte (parfois précède un rebond)."
            elif pos >= 0.65:
                boll_txt = "Le cours évolue dans la moitié haute des bandes de Bollinger → orientation positive."
            elif pos <= 0.35:
                boll_txt = "Le cours évolue dans la moitié basse des bandes de Bollinger → orientation négative."
            else:
                boll_txt = "Le cours évolue autour de sa moyenne mobile 20 jours (milieu des bandes) → pas de tension particulière."
            signals.append(("🎚️", "Volatilité (Bollinger 20,2)", boll_txt))

    # ── VERDICT GLOBAL ───────────────────────────────────────────────
    if score >= 2:
        bias = "haussier"
        verdict = "🟢 Configuration technique globalement HAUSSIÈRE : la majorité des indicateurs pointent dans le même sens positif."
    elif score <= -2:
        bias = "baissier"
        verdict = "🔴 Configuration technique globalement BAISSIÈRE : la majorité des indicateurs pointent dans le même sens négatif."
    else:
        bias = "neutre"
        verdict = "🟡 Configuration technique MITIGÉE / NEUTRE : les indicateurs s'équilibrent, pas de signal directionnel net."

    return {"verdict": verdict, "bias": bias, "bias_score": score, "signals": signals}


# ═════════════════════════════════════════════════════════════════════
#  HEADER
# ═════════════════════════════════════════════════════════════════════

st.markdown(f"""
<div class="bar-top">
    <span><span class="live-dot"></span>EN DIRECT · BRIEF MENSUEL EQUITY</span>
    <span>{datetime.now().strftime('%d/%m/%Y · %H:%M')}</span>
</div>
""", unsafe_allow_html=True)

st.markdown(f"# 📈 Brief Mensuel <span style='color:{COLORS['blue']}'>Bourse.</span>",
            unsafe_allow_html=True)
st.markdown(f"<p style='color:{COLORS['text_mid']}; font-size:15px;'>"
            f"+1200 actions analysées · PEA & CTO · Consensus analystes & dividendes · "
            f"Indices mondiaux · Analyse technique · Source Yahoo Finance"
            f"</p>", unsafe_allow_html=True)


# ═════════════════════════════════════════════════════════════════════
#  LOAD DATA
# ═════════════════════════════════════════════════════════════════════

sheets = load_xlsx_from_github()

if not sheets:
    st.warning(
        "⚠️ Aucune donnée disponible pour le moment. Le brief mensuel est publié le "
        "**premier jour ouvré de chaque mois**. Reviens bientôt !"
    )
    st.info(f"💼 [Suivre le Brief Mensuel sur LinkedIn]({LINKEDIN_URL})")
    st.stop()

df = sheets.get("By Total Gain", pd.DataFrame())
if df.empty:
    df = list(sheets.values())[0]  # fallback : 1er onglet dispo
if df.empty:
    st.error("❌ Données vides.")
    st.stop()

# Enrichissements colonnes
df["country"] = df["ticker"].apply(get_country_from_ticker)
if "reco_mean" in df.columns:
    df["stars"] = df["reco_mean"].apply(reco_to_stars)
if "reco_label" in df.columns:
    df["reco_label"] = (df["reco_label"].astype(str)
                        .replace({"none": "-", "None": "-", "nan": "-", "": "-"}))


# ═════════════════════════════════════════════════════════════════════
#  SIDEBAR - Filtres
# ═════════════════════════════════════════════════════════════════════

with st.sidebar:
    st.markdown("### 🔎 Recherche")
    search = st.text_input("Nom ou ticker", placeholder="ex: Apple, LVMH, MC.PA...").strip()

    st.markdown("### 🎯 Filtres")
    elig = st.radio("Éligibilité", ["Tout", "PEA uniquement", "CTO uniquement"])

    sectors_sel = []
    if "sector_fr" in df.columns:
        secteurs = sorted(df["sector_fr"].dropna().unique().tolist())
        sectors_sel = st.multiselect("Secteur GICS", secteurs, default=[])

    countries_sel = []
    countries = sorted(df["country"].unique().tolist())
    countries_sel = st.multiselect("Pays", countries, default=[])

    markets_sel = []
    if "market" in df.columns:
        markets = sorted(df["market"].dropna().unique().tolist())
        markets_sel = st.multiselect("Marché d'origine", markets, default=[])

    index_sel = []
    if "index_name" in df.columns:
        indices = sorted(df["index_name"].dropna().unique().tolist())
        index_sel = st.multiselect("Indice de provenance", indices, default=[],
                                   help="Indice boursier précis (CAC 40, DAX 40, Nikkei...)")

    CAP_BUCKETS = {
        "🐋 Méga cap (> 200 Md€)":   (200e9, float("inf")),
        "🐂 Grande cap (10–200 Md€)": (10e9, 200e9),
        "🐎 Moyenne cap (2–10 Md€)":  (2e9, 10e9),
        "🐭 Petite cap (< 2 Md€)":    (0, 2e9),
    }
    cap_sel = []
    if "market_cap_eur" in df.columns and df["market_cap_eur"].notna().any():
        cap_sel = st.multiselect("Capitalisation", list(CAP_BUCKETS.keys()), default=[],
                                 help="Filtre par taille d'entreprise (exclure les small caps, etc.)")

    st.markdown("### 📊 Métriques")
    pot_range = None
    if "total_pct" in df.columns:
        mn, mx = float(df["total_pct"].min()), float(df["total_pct"].max())
        pot_range = st.slider("Potentiel total (%)", mn, mx, (mn, mx), step=1.0, format="%.0f%%")

    perf_range = None
    if "perf_1m" in df.columns:
        dfp = df.dropna(subset=["perf_1m"])
        if not dfp.empty:
            mn, mx = float(dfp["perf_1m"].min()), float(dfp["perf_1m"].max())
            perf_range = st.slider("Performance mois (%)", mn, mx, (mn, mx), step=1.0, format="%.0f%%")

    div_min = 0.0
    if "div_pct" in df.columns:
        dmax = float(df["div_pct"].max())
        div_min = st.slider("Dividende minimum (%)", 0.0, max(1.0, dmax), 0.0, step=0.5, format="%.1f%%")

    price_range = None
    if "price_eur" in df.columns:
        dfpr = df.dropna(subset=["price_eur"])
        if not dfpr.empty:
            pmn, pmx = float(dfpr["price_eur"].min()), float(dfpr["price_eur"].max())
            price_range = st.slider("Cours (€)", pmn, pmx, (pmn, pmx), format="%.0f €")

    n_analysts_min = 0
    if "analyst_count" in df.columns:
        amax = int(df["analyst_count"].max()) if len(df) else 50
        n_analysts_min = st.slider("Couverture analystes (min)", 0, amax, 0)

    st.markdown("---")
    if st.button("🔄 Rafraîchir les données"):
        st.cache_data.clear()
        st.rerun()

    st.markdown("---")
    st.markdown("### 🔗 Liens")
    st.markdown(f"💼 [LinkedIn]({LINKEDIN_URL})")
    st.markdown(f"💳 [Parrainage Boursorama (+100€)]({PARRAINAGE_URL}) `{CODE_PARRAIN}`")


# ═════════════════════════════════════════════════════════════════════
#  APPLICATION DES FILTRES
# ═════════════════════════════════════════════════════════════════════

df_f = df.copy()

if search:
    s = search.lower()
    mask = (df_f["ticker"].str.lower().str.contains(s, na=False) |
            df_f["name"].str.lower().str.contains(s, na=False))
    df_f = df_f[mask]

if elig == "PEA uniquement":
    df_f = df_f[df_f["pea"] == True]
elif elig == "CTO uniquement":
    df_f = df_f[df_f["pea"] == False]

if sectors_sel:
    df_f = df_f[df_f["sector_fr"].isin(sectors_sel)]
if countries_sel:
    df_f = df_f[df_f["country"].isin(countries_sel)]
if markets_sel and "market" in df_f.columns:
    df_f = df_f[df_f["market"].isin(markets_sel)]
if index_sel and "index_name" in df_f.columns:
    df_f = df_f[df_f["index_name"].isin(index_sel)]
if cap_sel and "market_cap_eur" in df_f.columns:
    cap_mask = pd.Series(False, index=df_f.index)
    for label in cap_sel:
        lo, hi = CAP_BUCKETS[label]
        cap_mask |= df_f["market_cap_eur"].between(lo, hi)
    df_f = df_f[cap_mask]
if pot_range and "total_pct" in df_f.columns:
    df_f = df_f[df_f["total_pct"].between(pot_range[0], pot_range[1])]
if perf_range and "perf_1m" in df_f.columns:
    df_f = df_f[df_f["perf_1m"].between(perf_range[0], perf_range[1]) | df_f["perf_1m"].isna()]
if div_min > 0 and "div_pct" in df_f.columns:
    df_f = df_f[df_f["div_pct"] >= div_min]
if price_range and "price_eur" in df_f.columns:
    df_f = df_f[df_f["price_eur"].between(price_range[0], price_range[1]) | df_f["price_eur"].isna()]
if n_analysts_min > 0 and "analyst_count" in df_f.columns:
    df_f = df_f[df_f["analyst_count"] >= n_analysts_min]


# ═════════════════════════════════════════════════════════════════════
#  MÉTRIQUES GLOBALES
# ═════════════════════════════════════════════════════════════════════

n_total = len(df_f)
n_pea   = int(df_f["pea"].sum()) if "pea" in df_f.columns else 0
n_cto   = n_total - n_pea
avg_pot = df_f["total_pct"].mean() if n_total else 0
avg_div = df_f["div_pct"].mean() if (n_total and "div_pct" in df_f.columns) else 0

m1, m2, m3, m4, m5 = st.columns(5)
m1.metric("Actions affichées", f"{n_total}", delta=f"sur {len(df)} totales")
m2.metric("PEA · Zone EEE", f"{n_pea}", delta=f"{100*n_pea/max(1,n_total):.0f}%")
m3.metric("CTO · Mondial", f"{n_cto}", delta=f"{100*n_cto/max(1,n_total):.0f}%")
m4.metric("Potentiel moyen", fmt_pct(avg_pot))
m5.metric("Dividende moyen", fmt_pct(avg_div, plus=False))

st.markdown("---")


# ═════════════════════════════════════════════════════════════════════
#  TABS
# ═════════════════════════════════════════════════════════════════════

(tab_table, tab_indices, tab_charts, tab_ticker,
 tab_compare, tab_evol, tab_sectors, tab_alloc, tab_about) = st.tabs([
    "📋 Classement",
    "🌍 Indices",
    "📊 Graphiques",
    "🔍 Détail Ticker",
    "⚖️ Comparateur",
    "📈 Évolution",
    "📂 Par Secteur",
    "🎯 Allocation",
    "ℹ️ À propos",
])


# ─────────────────────────────────────────────────────────────────────
#  TAB 1 : CLASSEMENT COMPLET
# ─────────────────────────────────────────────────────────────────────
with tab_table:
    st.markdown(f"### Classement des {n_total} actions filtrées")

    sort_options = {
        "Potentiel total ↓ (cible + dividende)": ("total_pct", False),
        "Perf 1 mois ↓":              ("perf_1m", False),
        "Perf 1 mois ↑":              ("perf_1m", True),
        "Cible analystes ↓":          ("target_pct", False),
        "Dividende ↓":                ("div_pct", False),
        "Recommandation (meilleure)": ("reco_mean", True),
        "Couverture analystes ↓":     ("analyst_count", False),
        "Cours EUR ↑":                ("price_eur", True),
        "Cours EUR ↓":                ("price_eur", False),
        "Nom A→Z":                    ("name", True),
    }
    c_sort, c_info = st.columns([2, 3])
    with c_sort:
        sort_choice = st.selectbox("Trier par", list(sort_options.keys()))
    sort_col, sort_asc = sort_options[sort_choice]
    with c_info:
        st.caption("💡 Clique sur l'en-tête d'une colonne dans le tableau pour trier autrement. "
                   "Va dans l'onglet 🔍 Détail Ticker pour l'analyse complète d'une action.")

    df_show = df_f.copy()
    if sort_col in df_show.columns:
        df_show = df_show.sort_values(sort_col, ascending=sort_asc, na_position="last")
    df_show = df_show.reset_index(drop=True)
    df_show.insert(0, "Rang", range(1, len(df_show) + 1))

    # Colonne capi en Md€ (tri numérique + affichage propre)
    if "market_cap_eur" in df_show.columns:
        df_show["cap_mde"] = df_show["market_cap_eur"] / 1e9

    # Colonne Tendance technique (calculée dans le pipeline → colonne tech_signal).
    # Fallback sur le signal composite (trend_signal) si la colonne n'existe pas encore.
    if "tech_signal" in df_show.columns:
        df_show["signal"] = df_show["tech_signal"]
    else:
        df_show["signal"] = df_show.apply(lambda r: " ".join(trend_signal(r.to_dict())), axis=1)

    cols_show = [c for c in [
        "Rang", "ticker", "name", "stars", "signal", "country", "index_name", "sector_fr", "pea",
        "price_eur", "cap_mde", "perf_1m", "target_pct", "target_low_pct", "target_high_pct",
        "div_pct", "total_pct", "reco_label", "analyst_count",
        "boursorama_link", "yahoo_link",
    ] if c in df_show.columns]

    rename = {
        "ticker": "Ticker", "name": "Nom", "stars": "Note", "signal": "Tendance",
        "country": "Pays", "index_name": "Indice", "sector_fr": "Secteur", "pea": "PEA",
        "price_eur": "Cours €", "cap_mde": "Capi",
        "perf_1m": "Perf 1M", "target_pct": "Cible 12M",
        "target_low_pct": "Cible basse", "target_high_pct": "Cible haute",
        "div_pct": "Dividende", "total_pct": "Potentiel",
        "reco_label": "Conseil", "analyst_count": "Analystes",
        "boursorama_link": "Boursorama", "yahoo_link": "Yahoo",
    }
    df_disp = df_show[cols_show].rename(columns=rename)

    st.dataframe(
        df_disp, height=620, use_container_width=True, hide_index=True,
        column_config={
            "Note":        st.column_config.TextColumn("Note", help="Consensus analystes (★ = achat)"),
            "Tendance":    st.column_config.TextColumn("Tendance", help="Signal d'analyse technique pure (moyennes mobiles 20/50/200 + RSI + MACD), calculé sur 1 an de cours. 🟢 Haussier · 🟡 Neutre · 🔴 Baissier. Détail indicateur par indicateur dans l'onglet 🔍 Détail Ticker."),
            "Indice":      st.column_config.TextColumn("Indice", help="Indice boursier de provenance"),
            "PEA":         st.column_config.CheckboxColumn("PEA"),
            "Cours €":     st.column_config.NumberColumn("Cours €", format="%.2f €"),
            "Capi":        st.column_config.NumberColumn("Capi", format="%.1f Md€", help="Capitalisation boursière en milliards d'€"),
            "Perf 1M":     st.column_config.NumberColumn("Perf 1M", format="%+.1f %%"),
            "Cible 12M":   st.column_config.NumberColumn("Cible 12M", format="%+.1f %%"),
            "Cible basse": st.column_config.NumberColumn("Bas", format="%+.0f %%"),
            "Cible haute": st.column_config.NumberColumn("Haut", format="%+.0f %%"),
            "Dividende":   st.column_config.NumberColumn("Div", format="%.2f %%"),
            "Potentiel":   st.column_config.NumberColumn("Potentiel", format="%+.1f %%"),
            "Analystes":   st.column_config.NumberColumn("Analystes", format="%d"),
            "Boursorama":  st.column_config.LinkColumn("BR", display_text="🏛️"),
            "Yahoo":       st.column_config.LinkColumn("YF", display_text="🔍"),
        },
    )

    # ── Export (Excel multi-feuilles + CSV) ─────────────────────────
    # Colonnes export : noms FR lisibles, sans les liens bruts
    export_cols = [c for c in [
        "Rang", "ticker", "name", "stars", "country", "index_name", "sector_fr", "pea",
        "price_eur", "perf_1m", "target_pct", "target_low_pct", "target_high_pct",
        "div_pct", "total_pct", "reco_label", "analyst_count",
    ] if c in df_show.columns]

    def _prep_export(d: pd.DataFrame) -> pd.DataFrame:
        cols = [c for c in export_cols if c in d.columns]
        out = d[cols].rename(columns=rename)
        if "pea" in out.columns:  # au cas où non renommé
            out = out.rename(columns={"pea": "PEA"})
        if "PEA" in out.columns:
            out["PEA"] = out["PEA"].map({True: "Oui", False: "Non"})
        return out

    df_export = _prep_export(df_show)  # classement complet (pour CSV + feuille principale)

    pct_cols = ["Perf 1M", "Cible 12M", "Cible basse", "Cible haute", "Dividende", "Potentiel"]
    price_cols = ["Cours €"]

    # Feuilles du classeur
    sheets_xlsx = {
        "Classement complet": df_export,
        "Top PEA":  _prep_export(df_show[df_show["pea"] == True]) if "pea" in df_show.columns else None,
        "Top CTO":  _prep_export(df_show[df_show["pea"] == False]) if "pea" in df_show.columns else None,
    }
    # Feuille "Par secteur" : meilleur potentiel par secteur
    if "sector_fr" in df_show.columns:
        best_sec = (df_show.dropna(subset=["total_pct"])
                    .sort_values("total_pct", ascending=False)
                    .drop_duplicates(subset="sector_fr"))
        sheets_xlsx["Par secteur"] = _prep_export(best_sec)

    # Feuille Synthèse : KPIs + top 15
    kpis = {
        "Actions analysées :": int(n_total),
        "Éligibles PEA :": int(n_pea),
        "CTO (mondial) :": int(n_cto),
        "Potentiel moyen :": f"{avg_pot:+.1f} %",
        "Dividende moyen :": f"{avg_div:.2f} %",
        "Meilleur potentiel :": (f"{df_show.iloc[0]['name']} ({df_show.iloc[0]['total_pct']:+.1f} %)"
                                 if n_total and "total_pct" in df_show.columns else "-"),
    }
    top15 = _prep_export(df_show.nlargest(15, "total_pct")) if "total_pct" in df_show.columns else df_export.head(15)
    synth = {"kpis": kpis, "top_df": top15}

    c1, c2 = st.columns(2)
    with c1:
        xlsx_bytes = make_workbook_excel(sheets_xlsx, pct_cols, price_cols, synth=synth,
                                         color_scale_cols=["Potentiel"])
        st.download_button("📊 Télécharger Excel", xlsx_bytes,
                           file_name=f"brief_bourse_{datetime.now():%Y-%m-%d}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True,
                           help="Classeur multi-feuilles : Synthèse · Classement · Top PEA · Top CTO · Par secteur")
    with c2:
        csv = df_export.to_csv(index=False, sep=";").encode("utf-8-sig")  # ; + BOM = Excel FR friendly
        st.download_button("💾 Télécharger CSV", csv,
                           file_name=f"brief_bourse_{datetime.now():%Y-%m-%d}.csv",
                           mime="text/csv", use_container_width=True)


# ─────────────────────────────────────────────────────────────────────
#  TAB 2 : INDICES GLOBAUX
# ─────────────────────────────────────────────────────────────────────
with tab_indices:
    st.markdown("### 🌍 Indices & marchés mondiaux en temps réel")
    st.caption("Données live Yahoo Finance · perfs jour / mois / an · cache 30 min")

    idx_df = fetch_global_indices()
    if idx_df.empty:
        st.warning("⚠️ Impossible de charger les indices (rate-limit Yahoo). Réessaie dans quelques minutes.")
    else:
        # ── Tri ─────────────────────────────────────────────────────
        idx_sort_options = {
            "Performance 1 mois ↓": ("perf_month", False),
            "Performance jour ↓":   ("perf_day", False),
            "Performance 1 an ↓":   ("perf_year", False),
            "Performance 1 mois ↑": ("perf_month", True),
            "Ordre géographique":   (None, None),
        }
        idx_sort_choice = st.selectbox("Trier les indices par", list(idx_sort_options.keys()))
        idx_col, idx_asc = idx_sort_options[idx_sort_choice]
        if idx_col:
            idx_df = idx_df.sort_values(idx_col, ascending=idx_asc, na_position="last").reset_index(drop=True)

        # Cartes en grille 3 colonnes
        n_cols = 3
        rows_data = idx_df.to_dict("records")
        for i in range(0, len(rows_data), n_cols):
            cols = st.columns(n_cols)
            for j, col in enumerate(cols):
                if i + j >= len(rows_data):
                    break
                d = rows_data[i + j]
                with col:
                    perf_d = d["perf_day"]
                    arrow = "▲" if perf_d >= 0 else "▼"
                    clr = COLORS["green"] if perf_d >= 0 else COLORS["red"]
                    # mini sparkline
                    spark = go.Figure(go.Scatter(
                        y=d["series"], mode="lines",
                        line=dict(color=clr, width=1.5), showlegend=False,
                        hoverinfo="skip",
                    ))
                    spark.update_layout(
                        height=60, margin=dict(l=0, r=0, t=0, b=0),
                        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                        xaxis=dict(visible=False), yaxis=dict(visible=False),
                    )
                    st.markdown(
                        f"<div class='info-card'>"
                        f"<div class='label'>{d['label']}</div>"
                        f"<div class='value'>{d['last']:,.2f}</div>"
                        f"<div style='color:{clr}; font-weight:700; font-size:14px;'>"
                        f"{arrow} {perf_d:+.2f}% <span style='color:{COLORS['text_mid']}; font-size:11px;'>jour</span></div>"
                        f"</div>",
                        unsafe_allow_html=True,
                    )
                    st.plotly_chart(spark, use_container_width=True,
                                    config={"displayModeBar": False},
                                    key=f"spark_{d['ticker']}")
                    cc1, cc2 = st.columns(2)
                    cc1.markdown(f"<span style='color:{color_for_pct(d['perf_month'])}; font-size:12px;'>"
                                 f"1M {d['perf_month']:+.1f}%</span>", unsafe_allow_html=True)
                    cc2.markdown(f"<span style='color:{color_for_pct(d['perf_year'])}; font-size:12px;'>"
                                 f"1A {d['perf_year']:+.1f}%</span>", unsafe_allow_html=True)

        st.markdown("---")
        # Tableau récap triable
        st.markdown("#### 📋 Tableau récapitulatif")
        recap = idx_df[["label", "desc", "last", "perf_day", "perf_month", "perf_year"]].copy()
        recap.columns = ["Indice", "Description", "Dernier", "Jour", "1 Mois", "1 An"]
        st.dataframe(
            recap, use_container_width=True, hide_index=True,
            column_config={
                "Dernier": st.column_config.NumberColumn("Dernier", format="%.2f"),
                "Jour":    st.column_config.NumberColumn("Jour", format="%+.2f %%"),
                "1 Mois":  st.column_config.NumberColumn("1 Mois", format="%+.1f %%"),
                "1 An":    st.column_config.NumberColumn("1 An", format="%+.1f %%"),
            },
        )


# ─────────────────────────────────────────────────────────────────────
#  TAB 3 : GRAPHIQUES AGRÉGÉS
# ─────────────────────────────────────────────────────────────────────
with tab_charts:
    st.markdown("### 📊 Vue agrégée du classement filtré")
    if n_total < 2:
        st.info("Pas assez de données pour les graphiques (relâche les filtres).")
    else:
        st.markdown("#### 🎯 Cible analystes vs Performance mois")
        ds = df_f.dropna(subset=["target_pct", "perf_1m"])
        if not ds.empty:
            fig1 = px.scatter(
                ds, x="perf_1m", y="target_pct", color="sector_fr",
                hover_data=["ticker", "name", "country", "total_pct"],
                size="analyst_count" if "analyst_count" in ds.columns else None,
                size_max=28,
                labels={"perf_1m": "Performance mois (%)", "target_pct": "Cible analystes 12M (%)"},
            )
            fig1.add_hline(y=0, line_dash="dash", line_color=COLORS["dim"])
            fig1.add_vline(x=0, line_dash="dash", line_color=COLORS["dim"])
            st.plotly_chart(plot_styled_layout(fig1, height=500), use_container_width=True)

        ca, cb = st.columns(2)
        with ca:
            st.markdown("#### 📊 Distribution du potentiel")
            fig2 = px.histogram(df_f, x="total_pct", nbins=40,
                                color_discrete_sequence=[COLORS["blue"]],
                                labels={"total_pct": "Potentiel total (%)"})
            st.plotly_chart(plot_styled_layout(fig2, height=360), use_container_width=True)
        with cb:
            st.markdown("#### 📦 Performance par secteur")
            db = df_f.dropna(subset=["perf_1m"])
            if not db.empty:
                fig3 = px.box(db, y="sector_fr", x="perf_1m", color="sector_fr",
                              orientation="h", labels={"perf_1m": "Perf 1M (%)", "sector_fr": ""})
                fig3.update_layout(showlegend=False)
                st.plotly_chart(plot_styled_layout(fig3, height=360), use_container_width=True)

        st.markdown("#### 🌳 Treemap secteurs (taille = nb tickers, couleur = potentiel moyen)")
        dt = (df_f.groupby("sector_fr")
                  .agg(n=("ticker", "count"), avg_pot=("total_pct", "mean")).reset_index())
        fig4 = px.treemap(dt, path=["sector_fr"], values="n", color="avg_pot",
                          color_continuous_scale=[[0, COLORS["red"]], [0.5, COLORS["amber"]], [1, COLORS["green"]]],
                          color_continuous_midpoint=0,
                          labels={"avg_pot": "Potentiel moyen (%)"})
        st.plotly_chart(plot_styled_layout(fig4, height=450), use_container_width=True)

        st.markdown("#### 🏆 Top 20 par potentiel total")
        top20 = df_f.nlargest(20, "total_pct")
        fig5 = px.bar(top20.iloc[::-1], x="total_pct", y="ticker", color="total_pct",
                      color_continuous_scale=[[0, COLORS["red"]], [0.5, COLORS["amber"]], [1, COLORS["green"]]],
                      text="name", hover_data=["sector_fr", "country", "perf_1m"],
                      labels={"total_pct": "Potentiel total (%)", "ticker": ""})
        fig5.update_traces(textposition="inside")
        st.plotly_chart(plot_styled_layout(fig5, height=600), use_container_width=True)


# ─────────────────────────────────────────────────────────────────────
#  TAB 4 : DÉTAIL TICKER  (fondamental live + chart trading pro + Claude)
# ─────────────────────────────────────────────────────────────────────
with tab_ticker:
    st.markdown("### 🔍 Analyse complète d'une action")

    tickers_sorted = df_f.sort_values("total_pct", ascending=False, na_position="last")["ticker"].tolist()
    if not tickers_sorted:
        st.info("Pas de tickers dans la sélection actuelle (relâche les filtres).")
    else:
        c_sel, c_per = st.columns([3, 1])
        with c_sel:
            def _fmt_ticker(t):
                r = df_f[df_f["ticker"] == t].iloc[0]
                pot = r.get("total_pct")
                pot_s = f"{pot:+.1f}%" if pd.notna(pot) else "-"
                return f"{t} · {r['name']} · 🎯 {pot_s}"
            ticker_sel = st.selectbox(
                "Choisis une action (triée par potentiel décroissant)", tickers_sorted,
                format_func=_fmt_ticker,
            )
        with c_per:
            period_sel = st.selectbox(
                "Période", ["3mo", "6mo", "1y", "2y", "5y"], index=2,
                format_func=lambda p: {"3mo": "3 mois", "6mo": "6 mois", "1y": "1 an",
                                       "2y": "2 ans", "5y": "5 ans"}[p],
            )

        row = df_f[df_f["ticker"] == ticker_sel].iloc[0].to_dict()

        # En-tête
        st.markdown(f"#### {row.get('name', ticker_sel)} ({ticker_sel})")
        st.markdown(
            f"<span style='color:{COLORS['text_mid']}'>"
            f"{get_country_from_ticker(ticker_sel)} · {row.get('sector_fr', '-')} · "
            f"{'✅ Éligible PEA' if row.get('pea') else '🌍 CTO uniquement'} · "
            f"ISIN {row.get('isin') or '-'}</span>",
            unsafe_allow_html=True,
        )

        # Métriques classement (issues du brief)
        k1, k2, k3, k4, k5 = st.columns(5)
        k1.metric("Cours", f"{row.get('price_eur', 0):.2f} €",
                  delta=f"{row.get('price', 0):.2f} {row.get('currency', '')}")
        k2.metric("Perf 1M", fmt_pct(row.get("perf_1m")))
        k3.metric("Cible 12M", fmt_pct(row.get("target_pct")))
        k4.metric("Dividende", fmt_pct(row.get("div_pct"), plus=False))
        k5.metric("Potentiel", fmt_pct(row.get("total_pct")))

        st.markdown(
            f"**Consensus analystes** : {reco_to_stars(row.get('reco_mean'))} "
            f"· {row.get('reco_label', '-')} "
            f"· {int(row.get('analyst_count', 0) or 0)} analystes "
            f"· fourchette {fmt_pct(row.get('target_low_pct'))} → {fmt_pct(row.get('target_high_pct'))}"
        )
        with st.expander("❓ C'est quoi le « consensus analystes » et comment c'est calculé ?"):
            st.markdown(
                "Le **consensus** agrège l'avis des analystes financiers professionnels (banques, "
                "brokers) qui suivent l'action, via les données Yahoo Finance. Voici **comment chaque "
                "chiffre est calculé** :\n\n"
                "- **★ La note** : basée sur `recommendationMean` de Yahoo = la **moyenne des notes** "
                "des analystes, sur une échelle 1 à 5 (1 = achat fort, 5 = vente forte). On la convertit "
                "en étoiles avec la formule `6 − note` → une moyenne de 1 donne 5★, une moyenne de 5 "
                "donne 1★. Le libellé (« Achat fort », « Conserver »...) traduit cette moyenne.\n"
                "- **Cible 12M** : `(cible − cours actuel) / cours actuel × 100`. La « cible » est le "
                "**prix-objectif médian** des analystes à ~12 mois (`targetMedianPrice`). Ex : cours 100€, "
                "cible médiane 120€ → +20%.\n"
                "- **Fourchette** : même calcul mais avec le prix-objectif **le plus haut** "
                "(`targetHighPrice`, l'analyste le plus optimiste) et **le plus bas** (`targetLowPrice`, "
                "le plus pessimiste). Plus elle est large, plus les avis divergent.\n"
                "- **Dividende** : `dividende annuel / cours × 100` = le rendement du dividende.\n"
                "- **Potentiel total** : `Cible 12M + Dividende`. C'est le gain total espéré sur ~1 an "
                "(plus-value visée + dividende encaissé). C'est sur cette valeur qu'est fait le classement.\n"
                "- **Nombre d'analystes** (`numberOfAnalystOpinions`) : combien suivent le titre. Plus il "
                "y en a, plus le consensus est robuste. 1-2 analystes = à prendre avec prudence.\n\n"
                "⚠️ Le consensus est une **opinion de marché**, pas une vérité : les analystes se trompent "
                "souvent, surtout sur les petites valeurs et sur 12 mois."
            )

        # ── SIGNAL DE TENDANCE PRINCIPAL (technique, live) ──────────
        with st.spinner("Calcul du signal de tendance..."):
            hist_sig = fetch_ticker_history(ticker_sel, period=period_sel)
        if not hist_sig.empty and "Close" in hist_sig.columns:
            sig_sum = build_technical_summary(hist_sig)
            sig_color = {"haussier": COLORS["green"], "baissier": COLORS["red"],
                         "neutre": COLORS["amber"]}.get(sig_sum["bias"], COLORS["amber"])
            st.markdown(
                f"<div style='background:{COLORS['bg_panel']}; border:1px solid {sig_color}; "
                f"border-left:5px solid {sig_color}; padding:14px 20px; border-radius:6px; margin:10px 0;'>"
                f"<div style='color:{COLORS['text_mid']}; font-size:11px; text-transform:uppercase; "
                f"letter-spacing:1px;'>📡 Signal de tendance technique (sur {period_sel})</div>"
                f"<div style='font-size:18px; font-weight:800; color:{COLORS['text']}; margin-top:4px;'>"
                f"{sig_sum['verdict']}</div>"
                f"<div style='color:{COLORS['text_mid']}; font-size:12px; margin-top:6px;'>"
                f"Détail des indicateurs (RSI, MACD, moyennes, Bollinger) plus bas ⬇️</div>"
                f"</div>",
                unsafe_allow_html=True,
            )

        # Liens
        lc1, lc2, lc3 = st.columns(3)
        if row.get("boursorama_link"): lc1.markdown(f"🏛️ [Boursorama]({row['boursorama_link']})")
        if row.get("yahoo_link"):      lc2.markdown(f"🔍 [Yahoo Finance]({row['yahoo_link']})")
        if row.get("google_link"):     lc3.markdown(f"🔎 [Google Finance]({row['google_link']})")

        st.markdown("---")

        # ── FONDAMENTAL LIVE (yf.Ticker.info) ──────────────────────
        st.markdown("#### 🏢 Données fondamentales (live)")
        with st.spinner("Chargement du fondamental..."):
            info = fetch_ticker_info(ticker_sel)

        if info:
            f1, f2, f3, f4 = st.columns(4)
            f1.metric("Capitalisation", fmt_large_number(info.get("marketCap"), ""))
            pe = info.get("trailingPE")
            f2.metric("PER (trailing)", f"{pe:.1f}" if pe else "-")
            fpe = info.get("forwardPE")
            f3.metric("PER (forward)", f"{fpe:.1f}" if fpe else "-")
            beta = info.get("beta")
            f4.metric("Beta", f"{beta:.2f}" if beta else "-")

            g1, g2, g3, g4 = st.columns(4)
            pb = info.get("priceToBook")
            g1.metric("Price / Book", f"{pb:.2f}" if pb else "-")
            w52h = info.get("fiftyTwoWeekHigh")
            g2.metric("Haut 52 sem.", f"{w52h:.2f}" if w52h else "-")
            w52l = info.get("fiftyTwoWeekLow")
            g3.metric("Bas 52 sem.", f"{w52l:.2f}" if w52l else "-")
            emp = info.get("fullTimeEmployees")
            g4.metric("Employés", f"{emp:,}".replace(",", " ") if emp else "-")

            industry = info.get("industry")
            website = info.get("website")
            meta_line = []
            if industry: meta_line.append(f"**Industrie** : {industry}")
            if website:  meta_line.append(f"[🌐 Site web]({website})")
            if meta_line:
                st.markdown(" · ".join(meta_line))

            summary = info.get("longBusinessSummary")
            if summary:
                with st.expander("📖 Description de l'entreprise"):
                    st.write(summary)
        else:
            st.caption("ℹ️ Fondamental indisponible pour ce titre (ou rate-limit Yahoo).")

        st.markdown("---")

        # ── BOUTON CLAUDE ──────────────────────────────────────────
        prompt = build_claude_prompt(row, info)
        cb1, cb2 = st.columns([1, 2])
        with cb1:
            prompt_js = json.dumps(prompt)
            st.components.v1.html(f"""
            <div>
              <button onclick="copyAndRedirect()" style="
                background:#00bfe6;color:#0a1628;border:none;padding:12px 24px;
                font-weight:800;font-size:14px;letter-spacing:1px;cursor:pointer;
                border-radius:4px;width:100%;font-family:'JetBrains Mono',monospace;
                text-transform:uppercase;">🤖 Analyser avec Claude</button>
              <div id="status" style="margin-top:10px;color:#00ff66;font-family:monospace;font-size:13px;"></div>
            </div>
            <script>
            function copyAndRedirect() {{
                const p = {prompt_js};
                navigator.clipboard.writeText(p).then(() => {{
                    document.getElementById('status').innerHTML = '✅ Prompt copié ! Ouverture de Claude...';
                    setTimeout(() => window.open('https://claude.ai/new', '_blank'), 500);
                }}).catch(e => {{
                    document.getElementById('status').innerHTML = '⚠️ Copie manuelle nécessaire (voir ci-dessous).';
                }});
            }}
            </script>
            """, height=110)
        with cb2:
            st.info("🤖 Le clic copie un prompt d'analyse complet (chiffres + fondamental) dans "
                    "ton presse-papier et ouvre Claude.ai. Colle (Ctrl+V) pour lancer l'analyse.")
        with st.expander("📄 Voir le prompt"):
            st.code(prompt, language="markdown")

        st.markdown("---")

        # ── CHART TRADING PRO ──────────────────────────────────────
        st.markdown(f"#### 📈 Analyse technique sur {period_sel}")
        indicators = st.multiselect(
            "Indicateurs à afficher",
            ["Chandeliers", "Moyennes mobiles", "Bollinger", "Volume", "RSI", "MACD"],
            default=["Chandeliers", "Moyennes mobiles", "Bollinger", "Volume", "RSI"],
        )

        with st.spinner(f"Chargement de l'historique {ticker_sel}..."):
            hist = fetch_ticker_history(ticker_sel, period=period_sel)

        if hist.empty or "Close" not in hist.columns:
            st.warning(f"⚠️ Historique indisponible pour {ticker_sel}")
        else:
            close = hist["Close"]
            show_rsi = "RSI" in indicators
            show_macd = "MACD" in indicators
            show_vol = "Volume" in indicators and "Volume" in hist.columns

            # Construction des sous-graphes dynamiques
            rows_spec = [("price", 0.5)]
            if show_vol:  rows_spec.append(("vol", 0.15))
            if show_rsi:  rows_spec.append(("rsi", 0.17))
            if show_macd: rows_spec.append(("macd", 0.18))
            n_rows = len(rows_spec)
            heights = [h for _, h in rows_spec]
            total_h = sum(heights)
            heights = [h / total_h for h in heights]

            fig = make_subplots(
                rows=n_rows, cols=1, shared_xaxes=True,
                vertical_spacing=0.03, row_heights=heights,
            )
            row_idx = {name: i + 1 for i, (name, _) in enumerate(rows_spec)}

            # --- PRIX ---
            if "Chandeliers" in indicators and all(c in hist.columns for c in ["Open", "High", "Low", "Close"]):
                fig.add_trace(go.Candlestick(
                    x=hist.index, open=hist["Open"], high=hist["High"],
                    low=hist["Low"], close=hist["Close"], name="Cours",
                    increasing_line_color=COLORS["green"], decreasing_line_color=COLORS["red"],
                ), row=row_idx["price"], col=1)
            else:
                fig.add_trace(go.Scatter(
                    x=hist.index, y=close, name="Cours",
                    line=dict(color=COLORS["blue"], width=2),
                ), row=row_idx["price"], col=1)

            if "Moyennes mobiles" in indicators:
                for win, clr in [(20, COLORS["amber"]), (50, COLORS["text_mid"]), (200, COLORS["dim"])]:
                    ma = close.rolling(win).mean()
                    if ma.notna().any():
                        fig.add_trace(go.Scatter(
                            x=hist.index, y=ma, name=f"MA{win}",
                            line=dict(color=clr, width=1.3, dash="dot"),
                        ), row=row_idx["price"], col=1)

            if "Bollinger" in indicators:
                bmid, bup, blo = compute_bollinger(close)
                fig.add_trace(go.Scatter(x=hist.index, y=bup, name="Boll. sup",
                                         line=dict(color=COLORS["purple"], width=1), opacity=0.6),
                              row=row_idx["price"], col=1)
                fig.add_trace(go.Scatter(x=hist.index, y=blo, name="Boll. inf",
                                         line=dict(color=COLORS["purple"], width=1), opacity=0.6,
                                         fill="tonexty", fillcolor="rgba(180,124,255,0.08)"),
                              row=row_idx["price"], col=1)

            # --- VOLUME ---
            if show_vol:
                vol_colors = [COLORS["green"] if close.iloc[i] >= close.iloc[i-1] else COLORS["red"]
                              for i in range(1, len(close))]
                vol_colors = [COLORS["dim"]] + vol_colors
                fig.add_trace(go.Bar(x=hist.index, y=hist["Volume"], name="Volume",
                                     marker_color=vol_colors, showlegend=False),
                              row=row_idx["vol"], col=1)

            # --- RSI ---
            if show_rsi:
                rsi = compute_rsi(close)
                fig.add_trace(go.Scatter(x=hist.index, y=rsi, name="RSI",
                                         line=dict(color=COLORS["blue"], width=1.5)),
                              row=row_idx["rsi"], col=1)
                fig.add_hline(y=70, line_dash="dash", line_color=COLORS["red"],
                              opacity=0.5, row=row_idx["rsi"], col=1)
                fig.add_hline(y=30, line_dash="dash", line_color=COLORS["green"],
                              opacity=0.5, row=row_idx["rsi"], col=1)

            # --- MACD ---
            if show_macd:
                macd, sig, mhist = compute_macd(close)
                hist_colors = [COLORS["green"] if v >= 0 else COLORS["red"] for v in mhist]
                fig.add_trace(go.Bar(x=hist.index, y=mhist, name="Hist",
                                     marker_color=hist_colors, showlegend=False),
                              row=row_idx["macd"], col=1)
                fig.add_trace(go.Scatter(x=hist.index, y=macd, name="MACD",
                                         line=dict(color=COLORS["blue"], width=1.3)),
                              row=row_idx["macd"], col=1)
                fig.add_trace(go.Scatter(x=hist.index, y=sig, name="Signal",
                                         line=dict(color=COLORS["amber"], width=1.3)),
                              row=row_idx["macd"], col=1)

            fig.update_layout(
                height=300 + 130 * (n_rows - 1),
                paper_bgcolor=COLORS["bg"], plot_bgcolor=COLORS["bg_panel"],
                font=dict(family="JetBrains Mono, monospace", color=COLORS["text"], size=11),
                legend=dict(orientation="h", yanchor="bottom", y=1.01, xanchor="left", x=0,
                            bgcolor="rgba(0,0,0,0)"),
                margin=dict(l=40, r=20, t=30, b=20),
                xaxis_rangeslider_visible=False, hovermode="x unified",
            )
            for i in range(1, n_rows + 1):
                fig.update_xaxes(gridcolor=COLORS["grid"], row=i, col=1)
                fig.update_yaxes(gridcolor=COLORS["grid"], row=i, col=1)
            if show_rsi:
                fig.update_yaxes(range=[0, 100], row=row_idx["rsi"], col=1, title_text="RSI")
            if show_vol:
                fig.update_yaxes(title_text="Vol", row=row_idx["vol"], col=1)
            if show_macd:
                fig.update_yaxes(title_text="MACD", row=row_idx["macd"], col=1)

            st.plotly_chart(fig, use_container_width=True)

            # Stats résumé + interprétation
            s1, s2, s3, s4 = st.columns(4)
            perf_p = (close.iloc[-1] / close.iloc[0] - 1) * 100
            s1.metric("Perf période", fmt_pct(perf_p))
            rsi_now = compute_rsi(close).iloc[-1]
            if pd.notna(rsi_now):
                rsi_label = "Surachat" if rsi_now > 70 else "Survente" if rsi_now < 30 else "Neutre"
                s2.metric("RSI actuel", f"{rsi_now:.0f}", delta=rsi_label, delta_color="off")
            s3.metric("Plus haut", f"{close.max():.2f}")
            s4.metric("Plus bas", f"{close.min():.2f}")

            # ── LECTURE TECHNIQUE AUTOMATIQUE ───────────────────────
            st.markdown("#### 🧭 Lecture technique automatique")
            summary = build_technical_summary(hist)

            bias_color = {"haussier": COLORS["green"], "baissier": COLORS["red"],
                          "neutre": COLORS["amber"]}.get(summary["bias"], COLORS["amber"])
            st.markdown(
                f"<div style='background:{COLORS['bg_panel']}; border-left:4px solid {bias_color}; "
                f"padding:16px 20px; border-radius:4px; margin-bottom:14px;'>"
                f"<div style='font-size:17px; font-weight:700; color:{COLORS['text']};'>"
                f"{summary['verdict']}</div></div>",
                unsafe_allow_html=True,
            )

            for emoji, titre, texte in summary["signals"]:
                st.markdown(
                    f"<div style='padding:8px 0; border-bottom:1px solid {COLORS['grid']};'>"
                    f"<span style='color:{COLORS['blue']}; font-weight:700;'>{emoji} {titre}</span><br>"
                    f"<span style='color:{COLORS['text_mid']}; font-size:14px;'>{texte}</span></div>",
                    unsafe_allow_html=True,
                )

            st.caption("⚠️ Lecture automatique des indicateurs à titre informatif - ce n'est PAS un "
                       "conseil en investissement. Les indicateurs techniques décrivent le passé récent, "
                       "pas l'avenir. Risque de perte en capital.")


# ─────────────────────────────────────────────────────────────────────
#  TAB 5 : COMPARATEUR
# ─────────────────────────────────────────────────────────────────────
with tab_compare:
    st.markdown("### ⚖️ Comparateur d'actions (base 100)")
    st.caption("Sélectionne 2 à 4 actions pour comparer leur évolution normalisée.")

    all_tickers = df_f.sort_values("ticker")["ticker"].tolist()
    if len(all_tickers) < 2:
        st.info("Pas assez d'actions dans la sélection (relâche les filtres).")
    else:
        c_sel, c_per = st.columns([3, 1])
        with c_sel:
            cmp_sel = st.multiselect(
                "Actions à comparer (max 4)", all_tickers,
                default=all_tickers[:min(2, len(all_tickers))],
                max_selections=4,
                format_func=lambda t: f"{t} · {df_f[df_f['ticker']==t]['name'].iloc[0]}",
            )
        with c_per:
            cmp_period = st.selectbox("Période", ["3mo", "6mo", "1y", "2y", "5y"], index=2,
                                      format_func=lambda p: {"3mo": "3 mois", "6mo": "6 mois",
                                                            "1y": "1 an", "2y": "2 ans", "5y": "5 ans"}[p],
                                      key="cmp_period")

        if len(cmp_sel) >= 2:
            palette = [COLORS["blue"], COLORS["amber"], COLORS["green"], COLORS["purple"]]
            fig_cmp = go.Figure()
            rows_tab = []
            for i, tk in enumerate(cmp_sel):
                h = fetch_ticker_history(tk, period=cmp_period)
                if h.empty or "Close" not in h.columns:
                    continue
                close = h["Close"].dropna()
                base100 = close / close.iloc[0] * 100
                nm = df_f[df_f["ticker"] == tk]["name"].iloc[0]
                fig_cmp.add_trace(go.Scatter(
                    x=base100.index, y=base100, name=f"{tk}",
                    line=dict(color=palette[i % 4], width=2),
                ))
                r = df_f[df_f["ticker"] == tk].iloc[0].to_dict()
                rows_tab.append({
                    "Ticker": tk, "Nom": nm,
                    "Perf période": (close.iloc[-1] / close.iloc[0] - 1) * 100,
                    "Perf 1M": r.get("perf_1m"),
                    "Cible 12M": r.get("target_pct"),
                    "Dividende": r.get("div_pct"),
                    "Potentiel": r.get("total_pct"),
                })
            fig_cmp.add_hline(y=100, line_dash="dash", line_color=COLORS["dim"])
            fig_cmp.update_layout(
                yaxis_title="Base 100 au départ", xaxis_title="Date", hovermode="x unified")
            st.plotly_chart(plot_styled_layout(fig_cmp, height=480), use_container_width=True)

            if rows_tab:
                st.markdown("#### 📋 Comparaison chiffrée")
                cmp_df = pd.DataFrame(rows_tab)
                st.dataframe(
                    cmp_df, use_container_width=True, hide_index=True,
                    column_config={
                        "Perf période": st.column_config.NumberColumn("Perf période", format="%+.1f %%"),
                        "Perf 1M":      st.column_config.NumberColumn("Perf 1M", format="%+.1f %%"),
                        "Cible 12M":    st.column_config.NumberColumn("Cible 12M", format="%+.1f %%"),
                        "Dividende":    st.column_config.NumberColumn("Dividende", format="%.2f %%"),
                        "Potentiel":    st.column_config.NumberColumn("Potentiel", format="%+.1f %%"),
                    },
                )
        else:
            st.info("Sélectionne au moins 2 actions.")


# ─────────────────────────────────────────────────────────────────────
#  TAB 6 : ÉVOLUTION (vs mois dernier via snapshots)
# ─────────────────────────────────────────────────────────────────────
with tab_evol:
    st.markdown("### 📈 Évolution vs mois précédent")
    st.caption("Comparaison du classement actuel avec le snapshot du mois dernier (entrées, sorties, variations de rang).")

    now = datetime.now()
    cur_key = now.strftime("%Y-%m")
    prev_month = now.month - 1 or 12
    prev_year = now.year if now.month > 1 else now.year - 1
    prev_key = f"{prev_year:04d}-{prev_month:02d}"

    cc1, cc2 = st.columns(2)
    with cc1:
        cur_in = st.text_input("Mois courant (snapshot)", value=cur_key, help="Format AAAA-MM")
    with cc2:
        prev_in = st.text_input("Mois précédent (snapshot)", value=prev_key, help="Format AAAA-MM")

    snap_cur = load_snapshot(cur_in)
    snap_prev = load_snapshot(prev_in)

    if snap_cur.empty and snap_prev.empty:
        st.warning("⚠️ Aucun snapshot trouvé pour ces mois. Les snapshots s'accumulent au fil des mois - "
                   "reviens après plusieurs briefs pour voir l'évolution.")
    elif snap_prev.empty:
        st.info(f"📊 Snapshot {cur_in} trouvé ({len(snap_cur)} actions), mais pas {prev_in}. "
                "Pas encore de comparaison possible.")
    else:
        # Rang basé sur total_pct
        def add_rank(d):
            d = d.dropna(subset=["total_pct"]).sort_values("total_pct", ascending=False).reset_index(drop=True)
            d["rank"] = range(1, len(d) + 1)
            return d[["ticker", "name", "total_pct", "rank"]]

        rc = add_rank(snap_cur).set_index("ticker")
        rp = add_rank(snap_prev).set_index("ticker")

        cur_set, prev_set = set(rc.index), set(rp.index)
        entered = cur_set - prev_set
        exited = prev_set - cur_set
        common = cur_set & prev_set

        e1, e2, e3 = st.columns(3)
        e1.metric("Nouvelles entrées", len(entered))
        e2.metric("Sorties", len(exited))
        e3.metric("Maintenues", len(common))

        # Plus fortes progressions / chutes de rang
        moves = []
        for tk in common:
            delta = rp.loc[tk, "rank"] - rc.loc[tk, "rank"]  # positif = a monté
            moves.append({"Ticker": tk, "Nom": rc.loc[tk, "name"],
                          "Rang actuel": rc.loc[tk, "rank"], "Rang précédent": rp.loc[tk, "rank"],
                          "Évolution": delta})
        moves_df = pd.DataFrame(moves)

        if not moves_df.empty:
            cu, cd = st.columns(2)
            with cu:
                st.markdown("#### 🚀 Top progressions")
                top_up = moves_df.nlargest(10, "Évolution")
                st.dataframe(top_up, hide_index=True, use_container_width=True,
                             column_config={"Évolution": st.column_config.NumberColumn("Δ rang", format="%+d")})
            with cd:
                st.markdown("#### 📉 Top chutes")
                top_down = moves_df.nsmallest(10, "Évolution")
                st.dataframe(top_down, hide_index=True, use_container_width=True,
                             column_config={"Évolution": st.column_config.NumberColumn("Δ rang", format="%+d")})

        if entered:
            st.markdown("#### ✨ Nouvelles entrées dans le classement")
            ent_df = rc.loc[list(entered)].reset_index()[["ticker", "name", "total_pct", "rank"]]
            ent_df = ent_df.sort_values("rank").rename(
                columns={"ticker": "Ticker", "name": "Nom", "total_pct": "Potentiel", "rank": "Rang"})
            st.dataframe(ent_df, hide_index=True, use_container_width=True,
                         column_config={"Potentiel": st.column_config.NumberColumn("Potentiel", format="%+.1f %%")})


# ─────────────────────────────────────────────────────────────────────
#  TAB 7 : PAR SECTEUR
# ─────────────────────────────────────────────────────────────────────
with tab_sectors:
    st.markdown("### 📂 Top potentiels par secteur GICS")
    st.caption("Pour chaque secteur : meilleur ticker PEA vs meilleur ticker CTO (par potentiel total).")

    rows_al = []
    for sec in df_f["sector_fr"].dropna().unique():
        d = df_f[df_f["sector_fr"] == sec]
        bp = d[d["pea"] == True].nlargest(1, "total_pct")
        bc = d[d["pea"] == False].nlargest(1, "total_pct")
        rows_al.append({
            "Secteur": sec,
            "PEA": f"{bp.iloc[0]['name']} · {bp.iloc[0]['ticker']} ({bp.iloc[0]['total_pct']:+.1f}%)" if not bp.empty else "-",
            "PEA score": bp.iloc[0]["total_pct"] if not bp.empty else None,
            "CTO": f"{bc.iloc[0]['name']} · {bc.iloc[0]['ticker']} ({bc.iloc[0]['total_pct']:+.1f}%)" if not bc.empty else "-",
            "CTO score": bc.iloc[0]["total_pct"] if not bc.empty else None,
        })
    dal = pd.DataFrame(rows_al)
    if not dal.empty:
        dal["m"] = dal[["PEA score", "CTO score"]].max(axis=1)
        dal = dal.sort_values("m", ascending=False).drop(columns=["m"])
        st.dataframe(dal, use_container_width=True, hide_index=True, height=460,
                     column_config={
                         "PEA score": st.column_config.NumberColumn("PEA pot.", format="%+.1f %%"),
                         "CTO score": st.column_config.NumberColumn("CTO pot.", format="%+.1f %%"),
                     })

    st.markdown("---")
    st.markdown("### 📊 Statistiques par secteur")
    ss = (df_f.groupby("sector_fr").agg(
              Tickers=("ticker", "count"),
              Potentiel=("total_pct", "mean"),
              Perf=("perf_1m", "mean"),
              Dividende=("div_pct", "mean"),
          ).sort_values("Potentiel", ascending=False).reset_index()
          .rename(columns={"sector_fr": "Secteur"}))
    st.dataframe(ss, use_container_width=True, hide_index=True,
                 column_config={
                     "Potentiel": st.column_config.NumberColumn("Potentiel moy.", format="%+.1f %%"),
                     "Perf":      st.column_config.NumberColumn("Perf moy.", format="%+.1f %%"),
                     "Dividende": st.column_config.NumberColumn("Div moy.", format="%.2f %%"),
                 })


# ─────────────────────────────────────────────────────────────────────
#  TAB 8 : ALLOCATION (simulateur SOCLE / FUN)
# ─────────────────────────────────────────────────────────────────────
with tab_alloc:
    st.markdown("### 🎯 Simulateur d'allocation SOCLE / FUN")
    st.caption("Construis ton portefeuille selon la règle d'or : un SOCLE d'ETF mondiaux "
               "(sécurité, diversification) + une poche FUN de stock-picking (performance).")

    # Défaut = 1 action par secteur (nb de secteurs disponibles dans la sélection)
    n_secteurs = df_f["sector_fr"].nunique() if ("sector_fr" in df_f.columns and not df_f.empty) else 8
    n_secteurs = max(1, min(15, int(n_secteurs)))

    ac1, ac2, ac3 = st.columns(3)
    with ac1:
        capital = st.number_input("💰 Capital à investir (€)", min_value=100.0,
                                  value=10000.0, step=500.0, format="%.0f")
    with ac2:
        pct_socle = st.slider("🛡️ Part du SOCLE (ETF)", 30, 80, 55, step=5, format="%d%%")
    with ac3:
        n_fun = st.slider("🎲 Nombre de lignes FUN", 1, 15, n_secteurs,
                          help="Par défaut : 1 action par secteur (diversification maximale)")

    montant_socle = capital * pct_socle / 100
    montant_fun   = capital - montant_socle
    socle_sp      = montant_socle / 2
    socle_stoxx   = montant_socle / 2

    def _eur(x):
        return f"{x:,.0f} €".replace(",", " ")

    st.markdown("---")
    am1, am2, am3 = st.columns(3)
    am1.metric("🛡️ SOCLE (ETF)", _eur(montant_socle), delta=f"{pct_socle}%")
    am2.metric("🎲 FUN (actions)", _eur(montant_fun), delta=f"{100-pct_socle}%")
    am3.metric("💰 Total investi", _eur(capital))

    col_pie, col_txt = st.columns([1, 1])
    with col_pie:
        fig_alloc = go.Figure(go.Pie(
            labels=["🛡️ SOCLE · ETF S&P 500", "🛡️ SOCLE · ETF STOXX 600", "🎲 FUN · Stock-picking"],
            values=[socle_sp, socle_stoxx, montant_fun],
            marker=dict(colors=[COLORS["blue"], COLORS["blue_dim"], COLORS["amber"]]),
            hole=0.55, textinfo="percent",
        ))
        fig_alloc.update_layout(
            height=320, paper_bgcolor=COLORS["bg"], plot_bgcolor=COLORS["bg"],
            font=dict(family="JetBrains Mono, monospace", color=COLORS["text"], size=12),
            legend=dict(orientation="h", y=-0.1, bgcolor="rgba(0,0,0,0)"),
            margin=dict(l=10, r=10, t=10, b=10), showlegend=True,
        )
        st.plotly_chart(fig_alloc, use_container_width=True)
    with col_txt:
        st.markdown("#### 🛡️ Ton SOCLE")
        st.markdown(
            f"La base sécurisée : ~85% des fonds actifs sont battus par leur indice sur 10 ans "
            f"(étude SPIVA). On répartit en deux ETF mondiaux, 50/50 :\n\n"
            f"- **{_eur(socle_sp)}** sur un **ETF S&P 500** (actions US large cap)\n"
            f"- **{_eur(socle_stoxx)}** sur un **ETF STOXX 600** (actions européennes)\n\n"
            f"Astuce PEA : il existe des ETF S&P 500 et STOXX 600 éligibles PEA (à synthétique)."
        )

    st.markdown("---")
    st.markdown("#### 🎲 Ta poche FUN — suggestions depuis ton classement filtré")
    fun_elig = st.radio("Éligibilité de la poche FUN", ["Tout", "PEA uniquement", "CTO uniquement"],
                        horizontal=True)
    montant_par_ligne = montant_fun / n_fun if n_fun else 0
    st.caption(f"Soit **{_eur(montant_par_ligne)}** par ligne. Suggestions ci-dessous : "
               f"les meilleurs potentiels de ta sélection actuelle, **1 par secteur** (diversification).")

    df_fun = df_f.copy()
    if fun_elig == "PEA uniquement" and "pea" in df_fun.columns:
        df_fun = df_fun[df_fun["pea"] == True]
    elif fun_elig == "CTO uniquement" and "pea" in df_fun.columns:
        df_fun = df_fun[df_fun["pea"] == False]

    if df_fun.empty:
        st.info("Aucune action ne correspond (relâche les filtres ou change l'éligibilité).")
    else:
        fun_picks = (df_fun.dropna(subset=["total_pct"])
                     .sort_values("total_pct", ascending=False)
                     .drop_duplicates(subset="sector_fr")
                     .head(n_fun)
                     .copy())
        fun_picks["Montant"] = montant_par_ligne
        show_cols = [c for c in ["ticker", "name", "sector_fr", "country", "pea",
                                 "total_pct", "Montant"] if c in fun_picks.columns]
        fun_disp = fun_picks[show_cols].rename(columns={
            "ticker": "Ticker", "name": "Nom", "sector_fr": "Secteur",
            "country": "Pays", "pea": "PEA", "total_pct": "Potentiel",
        })
        st.dataframe(
            fun_disp, use_container_width=True, hide_index=True,
            column_config={
                "PEA":       st.column_config.CheckboxColumn("PEA"),
                "Potentiel": st.column_config.NumberColumn("Potentiel", format="%+.1f %%"),
                "Montant":   st.column_config.NumberColumn("Montant", format="%.0f €"),
            },
        )
        if len(fun_picks) < n_fun:
            st.caption(f"ℹ️ Seulement {len(fun_picks)} secteurs disponibles dans ta sélection "
                       f"(tu as demandé {n_fun} lignes). Élargis les filtres pour plus de diversité.")

    st.caption("⚠️ Simulateur pédagogique — ce n'est pas un conseil en investissement. "
               "Les montants sont indicatifs, à adapter à ta situation. Risque de perte en capital.")


# ─────────────────────────────────────────────────────────────────────
#  TAB 9 : À PROPOS
# ─────────────────────────────────────────────────────────────────────
with tab_about:
    st.markdown(f"""
### 📊 Brief Mensuel Bourse - Le produit d'analyse de marché

Application compagnon du **Brief Mensuel** publié chaque mois sur LinkedIn.
Explore librement les **+1200 actions** analysées : consensus analystes, dividendes,
indices mondiaux, analyse technique pro et analyse IA.

### 🎯 La règle d'or

> **SOCLE 50-60% = 2 ETF mondiaux** (S&P 500 + STOXX 600)
> **FUN 40-50% = stock-picking**, 1 action par secteur minimum
>
> Sur 10 ans, ~85% des fonds actifs sont battus par leur indice (étude SPIVA).
> Ton SOCLE c'est l'ETF, ton FUN c'est ce que tu trouves ici.

### 🧰 Fonctionnalités

- **Classement complet** : +1200 actions, tri/filtre, consensus, dividende, fourchette analystes, étoiles, ISIN, export CSV/Excel
- **Indices mondiaux** : ~15 indices live (S&P, NASDAQ, CAC, DAX, Nikkei, VIX, or, pétrole, Bitcoin...)
- **Détail ticker** : fondamental live (capitalisation, PER, beta, Price/Book...) + analyse technique (chandeliers, Bollinger, RSI, MACD, moyennes mobiles, volume)
- **Comparateur** : superpose 2-4 actions en base 100
- **Évolution** : entrées/sorties + variations de rang vs mois dernier
- **Analyse IA** : prompt structuré copié → Claude.ai en 1 clic

### 📡 Sources & méthode

- **Univers** : ~1500 actions (SP500 + STOXX 600 + DAX + MDAX + SBF120 + Nikkei + TSX60 + ASX50 + HSI...)
- **Filtre Boursorama** : seules les actions cotables sur Boursorama sont retenues
- **Données** : Yahoo Finance (consensus, cibles, dividendes, fondamental)
- **Conversion EUR** : taux FX Yahoo en temps réel
- **Mise à jour** : automatique le **premier jour ouvré** de chaque mois

### ⚠️ Disclaimer

Données issues de Yahoo Finance - **aucun conseil en investissement**. L'investissement en
bourse comporte un **risque de perte en capital**. Les performances passées ne préjugent pas
des performances futures. Fais tes propres recherches.

### 🔗 Liens

- 💼 LinkedIn : [Romain Taugourdeau]({LINKEDIN_URL})
- 💳 Parrainage Boursorama (+100€) : [{CODE_PARRAIN}]({PARRAINAGE_URL})

---
*Brief Mensuel Bourse · Données Yahoo Finance*
""")


# ═════════════════════════════════════════════════════════════════════
#  FOOTER
# ═════════════════════════════════════════════════════════════════════

st.markdown("---")
st.markdown(
    f"<div style='text-align:center; color:{COLORS['dim']}; font-size:12px; padding:20px 0;'>"
    f"<strong>BRIEF MENSUEL EQUITY</strong> · "
    f"« Risque de perte en capital. Ceci n'est pas un conseil en investissement. » · "
    f"Source : Yahoo Finance"
    f"</div>", unsafe_allow_html=True,
)
